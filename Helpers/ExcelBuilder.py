from openpyxl import load_workbook

class Menu:
    def __init__(self, file, restaurant):
        Menu_wb = load_workbook(filename = file)
        self.restaurant_ws = Menu_wb [restaurant]
        self.ORDERID = self.restaurant_ws['A']
        self.ITEMNAME = self.restaurant_ws['B']

    def list(self):
        menu = [self.ITEMNAME[i].value for i in range(0, len(self.ITEMNAME))]
        return menu

    def find_item(self, id):
        for row in self.restaurant_ws.iter_rows(min_row = 1):
            if(row[1].value == id):
                return [c.value for c in row]

    def item_stages(self, id):
        for row in self.restaurant_ws.iter_rows(min_row = 1):
            if(row[1].value == id):
                return [c.value for c in row] [2:]